"""
Extract a single blank security refund sheet with all formatting preserved
"""
import openpyxl
from copy import copy
import os
from datetime import datetime

def extract_single_sheet_template(source_file, output_file):
    """
    Extract a blank template with only one sheet, preserving all formatting
    """
    print(f"Loading source file: {source_file}")
    
    # Load the workbook
    wb = openpyxl.load_workbook(source_file)
    
    # Get the first sheet as template
    first_sheet_name = wb.sheetnames[0]
    print(f"Using sheet: {first_sheet_name}")
    ws = wb[first_sheet_name]
    
    # Find the last row with data
    max_row = ws.max_row
    
    # Clear only data rows, preserve headers, notes, and signature sections
    # Typically: Headers (1-15), Data rows (16-X), Notes/Signature (last 10-15 rows)
    data_end_row = max_row - 15  # Preserve last 15 rows for notes/signatures
    
    for row in ws.iter_rows(min_row=16, max_row=data_end_row):
        for cell in row:
            # Clear data entries only
            if cell.value and not isinstance(cell.value, str) or \
               (isinstance(cell.value, str) and cell.value and not cell.value.startswith('=')):
                cell.value = None
    
    print(f"Preserved header rows (1-15) and footer rows ({data_end_row+1}-{max_row}) with notes/signatures")
    
    # Remove all other sheets, keep only the first one
    sheets_to_remove = wb.sheetnames[1:]
    for sheet_name in sheets_to_remove:
        print(f"Removing sheet: {sheet_name}")
        wb.remove(wb[sheet_name])
    
    # Rename the sheet to a generic name
    ws.title = "Security Refund Sheet"
    
    # Save the blank template
    print(f"Saving blank template to: {output_file}")
    wb.save(output_file)
    print("Single sheet blank template created successfully!")

if __name__ == "__main__":
    # Source file from Output_Record
    source_file = "Output_Record/Excel_Files/output_17-09-2025_02-34/With_Deduction_fill_Batch_Full_01_17-09-2025.xlsx"
    
    # Output file
    output_file = "Blank_Security_Refund_Template.xlsx"
    
    if os.path.exists(source_file):
        extract_single_sheet_template(source_file, output_file)
        print(f"\n✓ Single sheet blank template created!")
        print(f"✓ Output file: {output_file}")
    else:
        print(f"Error: Source file not found: {source_file}")
