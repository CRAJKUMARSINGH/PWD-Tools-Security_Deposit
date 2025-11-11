"""
Simple PDF Export using HTML conversion
Alternative approach when Excel COM automation fails
"""

import os
import sys
import glob
from pathlib import Path
import openpyxl
import pandas as pd
import zipfile
from datetime import datetime
import weasyprint
from jinja2 import Template

def excel_to_html(excel_file, output_dir):
    """
    Convert Excel workbook sheets to HTML and then to PDF
    """
    try:
        print(f"Converting {excel_file}...")
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_file)
        converted_files = []
        
        base_name = Path(excel_file).stem
        
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                
                # Convert sheet to DataFrame
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                if not data:
                    continue
                
                # Create HTML content
                html_content = create_security_deposit_html(data, sheet_name)
                
                # Clean sheet name for filename
                safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
                pdf_filename = f"{base_name}_{safe_sheet_name}.pdf"
                pdf_path = os.path.join(output_dir, pdf_filename)
                
                # Convert HTML to PDF
                weasyprint.HTML(string=html_content).write_pdf(pdf_path)
                converted_files.append(pdf_path)
                print(f"  ✓ Exported sheet '{sheet_name}' to {pdf_filename}")
                
            except Exception as e:
                print(f"  ✗ Error exporting sheet '{sheet_name}': {e}")
        
        return converted_files
        
    except Exception as e:
        print(f"Error converting {excel_file}: {e}")
        return []

def create_security_deposit_html(data, sheet_name):
    """
    Create HTML representation of security deposit form
    """
    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>Security Deposit Refund - {{ sheet_name }}</title>
        <style>
            @page {
                size: A4;
                margin: 1cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 11px;
                line-height: 1.2;
                margin: 0;
                padding: 10px;
            }
            .title {
                text-align: center;
                font-weight: bold;
                font-size: 16px;
                color: #000080;
                background-color: #E6E6FA;
                padding: 8px;
                margin-bottom: 10px;
                border: 1px solid #000;
            }
            .form-field {
                margin: 3px 0;
                padding: 2px 0;
            }
            .field-label {
                font-weight: normal;
                display: inline-block;
                width: 60%;
            }
            .field-value {
                font-weight: bold;
                display: inline-block;
                width: 35%;
                border-bottom: 1px solid #000;
                min-height: 18px;
                padding-left: 5px;
            }
            .work-name {
                margin: 5px 0;
                padding: 5px;
                word-wrap: break-word;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin: 10px 0;
            }
            th, td {
                border: 1px solid #000;
                padding: 5px;
                text-align: center;
                font-size: 10px;
            }
            th {
                background-color: #E6E6FA;
                font-weight: bold;
            }
            .certification {
                margin-top: 15px;
                font-size: 9px;
            }
            .cert-title {
                font-weight: bold;
                margin-bottom: 5px;
            }
            .cert-item {
                margin: 2px 0;
                padding-left: 10px;
            }
            .signatures {
                margin-top: 20px;
                display: flex;
                justify-content: space-between;
            }
            .signature {
                text-align: center;
                width: 30%;
                border-top: 1px solid #000;
                padding-top: 5px;
                margin-top: 30px;
            }
        </style>
    </head>
    <body>
        {% for row_data in data %}
            {% if loop.index0 == 0 %}
                <div class="title">{{ row_data[0] if row_data[0] else "ORDER FOR REFUND OF SECURITY DEPOSIT [RWMF 119]" }}</div>
            {% elif row_data[0] and "Name of Work:" in (row_data[0] or "") %}
                <div class="work-name">{{ row_data[0] }}</div>
            {% elif row_data[0] and not row_data[0].startswith("Bill Num") and not row_data[0].startswith("Details") and not row_data[0].startswith("Certified") %}
                <div class="form-field">
                    <span class="field-label">{{ row_data[0] or "" }}</span>
                    <span class="field-value">{{ row_data[4] if row_data|length > 4 else (row_data[2] if row_data|length > 2 else "") }}</span>
                </div>
            {% endif %}
        {% endfor %}
        
        <!-- Security Deposit Table -->
        <div style="margin-top: 15px;">
            <div style="font-weight: bold; margin-bottom: 5px;">18. Details of Security Deposit</div>
            <table>
                <tr>
                    <th colspan="2">Bill Num</th>
                    <th>MB No.</th>
                    <th>Ded. Type</th>
                    <th>Amount (₹)</th>
                </tr>
                {% for i in range(6) %}
                <tr>
                    <td colspan="2">{% if i == 5 %}Total:{% endif %}</td>
                    <td></td>
                    <td></td>
                    <td></td>
                </tr>
                {% endfor %}
            </table>
        </div>
        
        <!-- Certification Section -->
        <div class="certification">
            <div class="cert-title">Certified That:-</div>
            <div class="cert-item">1. The Work has been completed as per G-schedule.</div>
            <div class="cert-item">2. The work has been inspected by the undersigned as on and it stood satisfactory.</div>
            <div class="cert-item">3. No Defect found during DLP Period.</div>
            <div class="cert-item">4. The final time extension granted upto With/without compensation by the competent authority.</div>
            <div class="cert-item">5. The defects pointed out by higher authorities or other authorized authorities during inspection etc have been removed by the contractor and compliance has been refund.</div>
        </div>
        
        <!-- Signatures -->
        <div class="signatures">
            <div class="signature">Divisional Accountant</div>
            <div class="signature">Assistant Engineer</div>
            <div class="signature">Executive Engineer<br>PWD Electric Div.- Udaipur</div>
        </div>
    </body>
    </html>
    """
    
    template = Template(html_template)
    return template.render(data=data, sheet_name=sheet_name)

def main():
    print("=" * 80)
    print("SIMPLE PDF EXPORT UTILITY")
    print("=" * 80)
    
    try:
        # Install weasyprint if not available
        import weasyprint
    except ImportError:
        print("Installing weasyprint...")
        subprocess.run([sys.executable, "-m", "pip", "install", "weasyprint"], check=True)
        import weasyprint
    
    # Find the latest output directory
    output_dirs = glob.glob("output_*")
    if not output_dirs:
        print("No output directories found. Please run the security deposit generator first.")
        return
    
    # Sort by modification time, newest first
    output_dirs.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    output_dir = output_dirs[0]
    
    print(f"Using latest output directory: {output_dir}")
    
    # Find all Excel files in the output directory
    excel_files = glob.glob(os.path.join(output_dir, "*.xlsx"))
    
    if not excel_files:
        print(f"No Excel files found in {output_dir}")
        return
    
    print(f"Found {len(excel_files)} Excel file(s) to convert:")
    for excel_file in excel_files:
        print(f"  - {os.path.basename(excel_file)}")
    
    # Create PDF output directory
    pdf_dir = f"PDF_Output_{os.path.basename(output_dir)}"
    
    print(f"\nConverting to PDF...")
    print(f"PDF output directory: {pdf_dir}")
    
    all_converted_files = []
    
    # Convert each Excel file
    for excel_file in excel_files:
        converted_files = excel_to_html(excel_file, pdf_dir)
        all_converted_files.extend(converted_files)
    
    print(f"\nConversion Summary:")
    print(f"  Total PDF files created: {len(all_converted_files)}")
    
    if all_converted_files:
        # Create ZIP archive
        print(f"\nCreating ZIP archive...")
        timestamp = datetime.now().strftime("%d%m%Y_%H%M")
        zip_name = f"PDF_Export_{timestamp}.zip"
        
        with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for pdf_file in all_converted_files:
                arcname = os.path.basename(pdf_file)
                zipf.write(pdf_file, arcname)
                print(f"  Added {arcname} to archive")
        
        print(f"✓ Created ZIP archive: {zip_name}")
        print(f"  Archive size: {os.path.getsize(zip_name) / 1024 / 1024:.1f} MB")
        
        print(f"\n✓ PDF Export completed successfully!")
        print(f"  PDF Directory: {pdf_dir}")
        print(f"  ZIP Archive: {zip_name}")
    else:
        print("No PDF files were created.")

if __name__ == "__main__":
    main()