"""
Enhanced Blank Security Refund Generator
Generates blank security deposit refund sheets with configurable file paths
PWD Electric Division - Udaipur
Developer: RAJKUMAR SINGH CHAUHAN
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.worksheet.hyperlink import Hyperlink
import os
import sys
from datetime import datetime, timedelta

class BlankSecurityRefundGenerator:
    """Class to handle blank security refund sheet generation"""
    
    def __init__(self, input_file=None):
        """Initialize with optional input file path"""
        self.input_file = input_file or self.find_input_file()
        self.output_dir = None
        
    def find_input_file(self):
        """Find the work order master file in current directory or parent directories"""
        possible_files = [
            "work_order_master.xlsx",
            os.path.join("..", "work_order_master.xlsx"),
            os.path.join("BLANK SD SHEETS", "work_order_master.xlsx"),
            os.path.join("..", "..", "work_order_master.xlsx")
        ]
        
        for file_path in possible_files:
            if os.path.exists(file_path):
                print(f"Found input file: {os.path.abspath(file_path)}")
                return os.path.abspath(file_path)
        
        # If not found, prompt user
        print("Work order master file not found in standard locations.")
        print("Please ensure 'work_order_master.xlsx' is in one of these locations:")
        for path in possible_files:
            print(f"  - {os.path.abspath(path)}")
        return None

    def read_excel_data(self, file_path, sheet_name='Work Orders'):
        """Read data from Excel file Work Orders sheet"""
        try:
            xl_file = pd.ExcelFile(file_path)
            print(f"Available sheets: {xl_file.sheet_names}")
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"Successfully read {len(df)} rows from {sheet_name} sheet")
            print(f"Columns: {list(df.columns)}")
            return df
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None

    def create_sheet_name(self, vendor, agreement_no):
        """Create sheet name from vendor and agreement number"""
        try:
            vendor_str = str(vendor).strip()
            vendor_clean = vendor_str.replace('M/s', '').replace('M/s. ', '').replace('M/s ', '').strip()
            first_name = vendor_clean.split()[0] if vendor_clean else 'Unknown'
            agreement_str = str(agreement_no).strip()
            agreement_clean = agreement_str.split('/')[0] if '/' in agreement_str else \
                             agreement_str.split('-')[0] if '-' in agreement_str else agreement_str
            sheet_name = f"{first_name} {agreement_clean}"
            invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
            for char in invalid_chars:
                sheet_name = sheet_name.replace(char, '')
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            if not sheet_name.strip():
                sheet_name = f"Work_{agreement_clean}"
            sheet_name = sheet_name.strip()
            return sheet_name
        except Exception as e:
            print(f"Error creating sheet name: {e}")
            return f"Work_{str(agreement_no).split('/')[0] if '/' in str(agreement_no) else str(agreement_no)}"

    def create_single_work_sheet(self, wb, row, work_idx):
        """Create a single work sheet with enhanced formatting"""
        # Get vendor name, agreement number, and work name from row
        vendor_name = row.get('Name of Contractor', '')
        agreement_no = row.get('Agreement No.', '')
        name_of_work = row.get('Name of Work', '')
        
        sheet_name = self.create_sheet_name(vendor_name, agreement_no or 'NoAgreement')
        ws = wb.create_sheet(title=sheet_name)

        # Define enhanced styles
        title_font = Font(bold=True, size=16, color='000080')
        header_font = Font(bold=True, size=12, color='000000')
        normal_font = Font(size=11, color='000000')
        small_font = Font(size=10, color='000000')
        value_font = Font(size=11, bold=True, color='000000')
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

        current_row = 1

        # Main title
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "ORDER FOR REFUND OF SECURITY DEPOSIT [RWMF 119]"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].fill = header_fill
        current_row += 1

        # Form fields - all blank for manual filling
        form_fields = [
            ("1. Name of Contractor:", ""),
            ("2. Amount of Deposit: ₹", ""),
            ("3. Name of Work:", ""),
            ("4. Agreement No.:", ""),
            ("5. Reference for granting refunds:", ""),
            ("6. Date of Commencement:", ""),
            ("7. Stipulated date of Completion:", ""),
            ("8. Actual Date of Completion:", ""),
            ("9. MB No.:", ""),
            ("10. Date of Payment of final bill:", ""),
            ("11. Date of Expiry of 3/6 months/DLP:", ""),
            ("12. Was work satisfactory:", ""),
            ("13. Any tools outstanding against contractor:", ""),
            ("14. Any recovery due from contractor after payment of final bill:", ""),
            ("15. Extension of time limit sanctioned vide", ""),
            ("16. Assistant Engineer Signature's Recommending refund", ""),
            ("17. Accountant's Remarks", "")
        ]

        for field_label, field_value in form_fields:
            field_num = int(field_label.split('.')[0]) if '.' in field_label else 0
            
            # Special handling for Name of Work (field 3)
            if field_num == 3:
                ws.merge_cells(f'A{current_row}:E{current_row}')
                ws[f'A{current_row}'] = field_label
                ws[f'A{current_row}'].font = normal_font
                ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            # Special handling for Name of Contractor (field 1)
            elif field_num == 1:
                ws[f'A{current_row}'] = field_label
                ws[f'A{current_row}'].font = normal_font
                ws[f'A{current_row}'].alignment = left_alignment
                ws.merge_cells(f'C{current_row}:E{current_row}')
                ws[f'C{current_row}'].border = thin_border
            # Special case for Amount of Deposit (field 2)
            elif field_num == 2:
                ws[f'A{current_row}'] = field_label
                ws[f'A{current_row}'].font = normal_font
                ws[f'A{current_row}'].alignment = left_alignment
                ws[f'E{current_row}'].border = thin_border
            # All other fields
            else:
                ws[f'A{current_row}'] = field_label
                ws[f'A{current_row}'].font = normal_font
                ws[f'A{current_row}'].alignment = left_alignment
                ws[f'E{current_row}'].border = thin_border
            current_row += 1

        # Security Deposit Details Table
        ws[f'A{current_row}'] = "18. Details of Security Deposit"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = left_alignment
        current_row += 1

        # Table headers
        headers = ["Bill Num", "MB No.", "Ded. Type", "Amount (₹)"]
        table_header_row = current_row
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = headers[0]
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].fill = header_fill
        
        for col_idx, header in enumerate(headers[1:], 1):
            col_letter = chr(67 + col_idx - 1)  # C, D, E
            ws[f'{col_letter}{current_row}'] = header
            ws[f'{col_letter}{current_row}'].font = header_font
            ws[f'{col_letter}{current_row}'].alignment = center_alignment
            ws[f'{col_letter}{current_row}'].border = thin_border
            ws[f'{col_letter}{current_row}'].fill = header_fill
        current_row += 1

        # Table data rows (blank for manual filling)
        table_data = [("", "", "", ""), ("", "", "", ""), ("", "", "", ""), 
                     ("", "", "", ""), ("", "", "", ""), ("Total:", "", "", "")]
        
        for table_row in table_data:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'] = table_row[0]
            ws[f'A{current_row}'].font = normal_font
            ws[f'A{current_row}'].alignment = center_alignment
            ws[f'A{current_row}'].border = thin_border
            
            for col_idx, value in enumerate(table_row[1:], 1):
                col_letter = chr(67 + col_idx - 1)  # C, D, E
                ws[f'{col_letter}{current_row}'] = value
                ws[f'{col_letter}{current_row}'].font = normal_font
                ws[f'{col_letter}{current_row}'].alignment = center_alignment
                ws[f'{col_letter}{current_row}'].border = thin_border
            current_row += 1

        # Certification section
        certification_items = [
            "Certified That:-",
            "1. The Work has been completed as per G-schedule.",
            "2. The work has been inspected by the undersigned as on and it stood satisfactory.",
            "3. No Defect found during DLP Period.",
            "4. The final time extension granted upto With/without compensation by the competent authority.",
            "5. The defects pointed out by higher authorities or other authorized authorities during inspection etc have been removed by the contractor and compliance has been refund."
        ]
        
        for cert_item in certification_items:
            if cert_item.startswith("Certified That:-"):
                ws[f'A{current_row}'] = cert_item
                ws[f'A{current_row}'].font = header_font
                ws[f'A{current_row}'].alignment = left_alignment
            elif cert_item.startswith("5."):
                ws.merge_cells(f'A{current_row}:E{current_row}')
                ws[f'A{current_row}'] = cert_item
                ws[f'A{current_row}'].font = small_font
                ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.row_dimensions[current_row].height = 40
            else:
                ws[f'A{current_row}'] = cert_item
                ws[f'A{current_row}'].font = small_font
                ws[f'A{current_row}'].alignment = left_alignment
            current_row += 1

        current_row += 1
        # Signature section
        signature_items = [
            ("Divisional Accountant", "Assistant Engineer", "Executive Engineer"),
            ("", "", "PWD Electric Div.- Udaipur")
        ]
        
        for sig_row in signature_items:
            for col_idx, sig_text in enumerate(sig_row):
                col_letter = chr(65 + col_idx * 2)  # A, C, E
                if sig_text:
                    ws[f'{col_letter}{current_row}'] = sig_text
                    ws[f'{col_letter}{current_row}'].font = normal_font
                    ws[f'{col_letter}{current_row}'].alignment = center_alignment
            current_row += 1

        # Set column widths and row heights
        for col, width in {'A': 30, 'B': 5, 'C': 25, 'D': 25, 'E': 25}.items():
            ws.column_dimensions[col].width = width
            
        for r in range(1, current_row + 5):
            ws.row_dimensions[r].height = 20

        # Setup print layout
        self.setup_print_layout(ws)
        return ws

    def setup_print_layout(self, ws):
        """Setup print layout for A4 portrait"""
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = 'portrait'
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True

    def create_security_refund_sheet(self, data_batch, batch_number, agreement_year=None):
        """Create security refund sheet for a batch of data"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for work_idx, (_, row) in enumerate(data_batch.iterrows(), 1):
            self.create_single_work_sheet(wb, row, work_idx)
        
        return wb

    def split_data_into_batches(self, df, batch_size=25):
        """Split dataframe into batches"""
        batches = []
        total_rows = len(df)
        for i in range(0, total_rows, batch_size):
            batch = df.iloc[i:i+batch_size].copy()
            batches.append((batch, (i // batch_size) + 1))
        return batches

    def get_agreement_year_from_data(self, df):
        """Extract agreement year from data"""
        try:
            if 'Agreement No' in df.columns:
                sample_agreement = str(df['Agreement No'].iloc[0]) if not df.empty else ''
                if len(sample_agreement) >= 4:
                    for i in range(len(sample_agreement) - 3):
                        year_candidate = sample_agreement[i:i+4]
                        if year_candidate.isdigit() and 2000 <= int(year_candidate) <= 2030:
                            return year_candidate
            return datetime.now().strftime('%Y')
        except Exception:
            return datetime.now().strftime('%Y')

    def generate_blank_sheets(self):
        """Main method to generate blank security refund sheets"""
        if not self.input_file or not os.path.exists(self.input_file):
            print(f"Error: Input file '{self.input_file}' not found.")
            print("Please ensure work_order_master.xlsx is available.")
            return False

        print(f"Reading Excel file from: {self.input_file}")
        df = self.read_excel_data(self.input_file, 'Work Orders')
        
        if df is None:
            print("Failed to read Excel file. Please check the file path and sheet name.")
            return False

        print(f"Total works found: {len(df)}")
        agreement_year = self.get_agreement_year_from_data(df)
        print(f"Using agreement year: {agreement_year}")

        batches = self.split_data_into_batches(df, 25)
        print(f"Created {len(batches)} batches")

        # Create output directory
        self.output_dir = f"BLANK_SD_SHEETS_{datetime.now().strftime('%d-%m-%Y_%H-%M')}"
        os.makedirs(self.output_dir, exist_ok=True)

        for batch_idx, (batch_data, batch_number) in enumerate(batches, 1):
            print(f"Processing batch {batch_idx} with {len(batch_data)} works...")
            wb = self.create_security_refund_sheet(batch_data, batch_idx, agreement_year)
            filename = f"Blank_Security_Refund_Batch_{batch_idx:02d}_{agreement_year}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            print(f"Saved: {filepath}")

        print(f"\n" + "="*80)
        print("BLANK SECURITY REFUND SHEETS GENERATION COMPLETED!")
        print("="*80)
        print(f"Generated {len(batches)} blank workbooks in '{self.output_dir}' directory.")
        print("Each workbook contains:")
        print("- 25 separate blank sheets (one per work order)")
        print("- Sheet names: Based on contractor names and agreement numbers") 
        print("- Professional formatting with proper borders")
        print("- Print-ready A4 portrait format, one page per sheet")
        print("- Ready for manual data entry")
        print("="*80)
        
        return True

def main():
    """Main function with improved path handling"""
    print("="*80)
    print("BLANK SECURITY REFUND GENERATOR")
    print("="*80)
    print("PWD Electric Division - Udaipur")
    print("Developer: RAJKUMAR SINGH CHAUHAN")
    print("="*80)
    
    # Initialize generator
    generator = BlankSecurityRefundGenerator()
    
    # Generate blank sheets
    success = generator.generate_blank_sheets()
    
    if success:
        print(f"\n✅ SUCCESS: Blank sheets generated in '{generator.output_dir}'")
    else:
        print(f"\n❌ FAILED: Could not generate blank sheets")
        sys.exit(1)

if __name__ == "__main__":
    main()