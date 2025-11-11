"""
Extract a blank security refund sheet template with all formatting preserved
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import os
from datetime import datetime

def extract_blank_template(source_file, output_file):
    """
    Extract a blank template from an existing security refund sheet
    Preserves all formatting including fonts, colors, borders, alignment, column widths, row heights
    """
    print(f"Loading source file: {source_file}")
    
    # Load the workbook
    wb = openpyxl.load_workbook(source_file)
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Find the data range (rows with actual data to clear)
        # Typically, security refund sheets have headers in first few rows
        # and data starts after that
        
        # Clear cell values while preserving formatting
        for row in ws.iter_rows():
            for cell in row:
                # Keep formulas and headers, clear data entries
                # Identify data rows (typically rows after row 10 or so)
                if cell.row > 15:  # Adjust this based on your sheet structure
                    if cell.value and not isinstance(cell.value, str) or \
                       (isinstance(cell.value, str) and cell.value and not cell.value.startswith('=')):
                        # Clear the value but keep the formatting
                        cell.value = None
    
    # Save the blank template
    print(f"Saving blank template to: {output_file}")
    wb.save(output_file)
    print("Blank template created successfully!")

if __name__ == "__main__":
    # Source file from Output_Record
    source_file = "Output_Record/Excel_Files/output_17-09-2025_02-34/With_Deduction_fill_Batch_Full_01_17-09-2025.xlsx"
    
    # Output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"Blank_Security_Refund_Template_{timestamp}.xlsx"
    
    if os.path.exists(source_file):
        extract_blank_template(source_file, output_file)
        print(f"\n✓ Blank template extracted successfully!")
        print(f"✓ Output file: {output_file}")
    else:
        print(f"Error: Source file not found: {source_file}")
